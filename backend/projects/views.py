"""
Views for Project API endpoints.
Implements all business logic from Flask app with role-based access control.
"""
from rest_framework import status, generics, filters
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db.models import Q
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
import csv
from datetime import datetime
from io import StringIO
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter

from .models import Project, LookupData
from .serializers import (
    ProjectSerializer, 
    ProjectCreateUpdateSerializer,
    LookupDataSerializer,
    BulkDeleteSerializer
)
from accounts.permissions import IsAdmin, IsOwnerOrAdmin
from .filters import ProjectFilter


class MyProjectsListView(generics.ListAPIView):
    """
    Get user's own projects (My Projects tab).
    Excludes soft-deleted projects.
    """
    serializer_class = ProjectSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = ProjectFilter
    search_fields = ['application_number', 'account_name', 'project_court', 'reviewed_by']
    ordering_fields = ['created_at', 'completed_date', 'application_number']
    ordering = ['-created_at']
    
    def get_queryset(self):
        """Return only user's own non-deleted projects."""
        return Project.objects.filter(
            created_by=self.request.user,
            is_deleted=False
        ).select_related('created_by')


class TeamProjectsListView(generics.ListAPIView):
    """
    Get all team projects (Team Projects tab).
    Admin: all projects
    User: team projects only
    Excludes soft-deleted projects.
    """
    serializer_class = ProjectSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = ProjectFilter
    search_fields = ['application_number', 'account_name', 'project_court', 'reviewed_by']
    ordering_fields = ['created_at', 'completed_date', 'application_number']
    ordering = ['-created_at']
    
    def get_queryset(self):
        """Return team projects based on user role."""
        user = self.request.user
        queryset = Project.objects.filter(is_deleted=False).select_related('created_by')
        
        if user.role == 'admin':
            # Admin sees all projects
            return queryset
        else:
            # Regular user sees projects from their team
            if user.team:
                return queryset.filter(created_by__team=user.team)
            else:
                # If user has no team, show only their own projects
                return queryset.filter(created_by=user)


class ProjectDetailView(generics.RetrieveAPIView):
    """
    Get project details.
    Users can view their own projects and team projects.
    Admins can view all projects.
    """
    serializer_class = ProjectSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """Filter based on user role and ownership."""
        user = self.request.user
        if user.is_admin:
            return Project.objects.filter(is_deleted=False)
        
        # Users can view own projects or team projects
        return Project.objects.filter(
            Q(created_by=user) | Q(created_by__team=user.team),
            is_deleted=False
        )


class ProjectCreateView(generics.CreateAPIView):
    """
    Create a new project.
    Automatically sets created_by to current user.
    """
    serializer_class = ProjectCreateUpdateSerializer
    permission_classes = [IsAuthenticated]
    
    def perform_create(self, serializer):
        """Set created_by to current user."""
        serializer.save(created_by=self.request.user)


class ProjectUpdateView(generics.UpdateAPIView):
    """
    Update a project.
    Users can only update their own drafts.
    Admins can update any project.
    """
    serializer_class = ProjectCreateUpdateSerializer
    permission_classes = [IsAuthenticated, IsOwnerOrAdmin]
    
    def get_queryset(self):
        """Filter based on user role and ownership."""
        user = self.request.user
        if user.is_admin:
            return Project.objects.filter(is_deleted=False)
        
        # Users can only update their own drafts
        return Project.objects.filter(
            created_by=user,
            stage='Started',
            is_deleted=False
        )


class ProjectDeleteView(generics.DestroyAPIView):
    """
    Soft delete a project.
    Users can only delete their own projects.
    Admins can delete any project.
    """
    permission_classes = [IsAuthenticated, IsOwnerOrAdmin]
    
    def get_queryset(self):
        """Filter based on user role and ownership."""
        user = self.request.user
        if user.is_admin:
            return Project.objects.filter(is_deleted=False)
        
        # Users can only delete their own projects
        return Project.objects.filter(created_by=user, is_deleted=False)
    
    def perform_destroy(self, instance):
        """Soft delete instead of hard delete."""
        instance.soft_delete(self.request.user)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def save_draft_view(request):
    """
    Save project as draft (stage='Started' for new, preserve for existing).
    Minimal validation - only required fields.
    Admin can edit any project, users can only edit their own.
    """
    data = request.data.copy()
    
    # Check if updating existing draft
    project_id = data.get('id')
    
    # If no stage provided and it's a new project, set to 'Started'
    # If updating existing project, preserve the existing stage
    if project_id:
        try:
            # Admin can edit any project, users can only edit their own
            if request.user.is_admin:
                project = Project.objects.get(id=project_id, is_deleted=False)
            else:
                project = Project.objects.get(id=project_id, created_by=request.user, is_deleted=False)
            
            # Preserve the existing stage if not provided in data
            if 'stage' not in data or not data['stage']:
                data['stage'] = project.stage
                
            serializer = ProjectCreateUpdateSerializer(project, data=data, partial=True)
        except Project.DoesNotExist:
            return Response(
                {'error': 'Project not found or you do not have permission to edit it.'},
                status=status.HTTP_404_NOT_FOUND
            )
    else:
        # New project - default to Started
        if 'stage' not in data or not data['stage']:
            data['stage'] = 'Started'
        serializer = ProjectCreateUpdateSerializer(data=data)
    
    if serializer.is_valid():
        if project_id:
            serializer.save()
        else:
            serializer.save(created_by=request.user)
        
        # Return full project data
        project = Project.objects.get(id=serializer.instance.id)
        return Response(
            ProjectSerializer(project).data,
            status=status.HTTP_200_OK if project_id else status.HTTP_201_CREATED
        )
    
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def submit_project_view(request):
    """
    Submit project (stage='Completed').
    Full validation - all required fields must be present.
    Auto-sets: completed_date (today), end_time (now in MST/IST).
    """
    from django.utils import timezone as tz
    import pytz
    
    data = request.data.copy()
    data['stage'] = 'Completed'
    
    # Auto-set completed_date to today (in MST)
    mst_tz = pytz.timezone('America/Phoenix')
    now_mst = tz.now().astimezone(mst_tz)
    data['completed_date'] = now_mst.date().isoformat()
    
    # Auto-set end_time to current time (will be converted to IST by serializer)
    data['end_time'] = now_mst.isoformat()
    
    # Check if updating existing project
    project_id = data.get('id')
    if project_id:
        try:
            project = Project.objects.get(id=project_id)
            
            # Permission check
            if not (request.user.is_admin or project.created_by == request.user):
                return Response(
                    {'error': 'You do not have permission to edit this project.'},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            serializer = ProjectCreateUpdateSerializer(project, data=data)
        except Project.DoesNotExist:
            return Response(
                {'error': 'Project not found.'},
                status=status.HTTP_404_NOT_FOUND
            )
    else:
        serializer = ProjectCreateUpdateSerializer(data=data)
    
    if serializer.is_valid():
        if project_id:
            serializer.save()
        else:
            serializer.save(created_by=request.user)
        
        # Return full project data
        project = Project.objects.get(id=serializer.instance.id)
        return Response(
            ProjectSerializer(project).data,
            status=status.HTTP_200_OK if project_id else status.HTTP_201_CREATED
        )
    
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdmin])
def bulk_delete_view(request):
    """
    Bulk delete projects (admin only).
    Soft deletes all specified projects.
    """
    serializer = BulkDeleteSerializer(data=request.data)
    if serializer.is_valid():
        project_ids = serializer.validated_data['project_ids']
        
        # Get projects to delete
        projects = Project.objects.filter(id__in=project_ids, is_deleted=False)
        
        # Soft delete all
        deleted_count = 0
        for project in projects:
            project.soft_delete(request.user)
            deleted_count += 1
        
        return Response({
            'message': f'Successfully deleted {deleted_count} project(s).',
            'deleted_count': deleted_count
        }, status=status.HTTP_200_OK)
    
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def lookup_data_view(request):
    """
    Get lookup data for dropdowns (courts, reviewers, etc.).
    Returns all active lookup data.
    """
    courts = LookupData.objects.filter(lookup_type='court', is_active=True).values_list('value', flat=True)
    reviewers = LookupData.objects.filter(lookup_type='reviewer', is_active=True).values_list('value', flat=True)
    
    return Response({
        'courts': list(courts),
        'reviewers': list(reviewers),
        'statuses': [choice[0] for choice in Project.STATUS_CHOICES],
        'third_party_options': [choice[0] for choice in Project.THIRD_PARTY_CHOICES],
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def filter_options_view(request):
    """
    Get filter options for My Projects.
    Returns unique values for dropdowns based on user's projects.
    """
    user_projects = Project.objects.filter(created_by=request.user, is_deleted=False)
    
    return Response({
        'courts': list(user_projects.values_list('project_court', flat=True).distinct().order_by('project_court')),
        'reviewers': list(user_projects.values_list('reviewed_by', flat=True).distinct().order_by('reviewed_by')),
        'statuses': list(user_projects.values_list('project_status', flat=True).distinct().order_by('project_status')),
        'stages': list(user_projects.values_list('stage', flat=True).distinct().order_by('stage')),
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def team_filter_options_view(request):
    """
    Get filter options for Team Projects.
    Admin: all projects
    User: team projects
    """
    user = request.user
    if user.is_admin:
        team_projects = Project.objects.filter(is_deleted=False)
    else:
        team_projects = Project.objects.filter(created_by__team=user.team, is_deleted=False)
    
    # Get unique created_by users
    creators = team_projects.values('created_by__id', 'created_by__username').distinct()
    
    return Response({
        'courts': list(team_projects.values_list('project_court', flat=True).distinct().order_by('project_court')),
        'reviewers': list(team_projects.values_list('reviewed_by', flat=True).distinct().order_by('reviewed_by')),
        'statuses': list(team_projects.values_list('project_status', flat=True).distinct().order_by('project_status')),
        'stages': list(team_projects.values_list('stage', flat=True).distinct().order_by('stage')),
        'creators': [{'id': c['created_by__id'], 'username': c['created_by__username']} for c in creators],
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdmin])
def export_excel_view(request):
    """
    Export projects to Excel (admin only).
    Supports date range filtering and exports all columns in MST timezone.
    """
    # Get filter parameters
    start_date = request.data.get('start_date')
    end_date = request.data.get('end_date')
    project_ids = request.data.get('project_ids', [])
    
    # Build queryset
    queryset = Project.objects.filter(is_deleted=False).select_related('created_by')
    
    if start_date:
        queryset = queryset.filter(completed_date__gte=start_date)
    if end_date:
        queryset = queryset.filter(completed_date__lte=end_date)
    if project_ids:
        queryset = queryset.filter(id__in=project_ids)
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Projects'
    
    # Headers
    headers = [
        'Application #', 'Account Name', 'Project Court', 'Reviewed By',
        'Status', 'Stage', 'Completed Date', 'Start Time (MST)', 'End Time (MST)',
        'Total Time (hrs)', 'Partner Installer Account', 'Third Party Salesforce',
        'Comments', 'Content', 'Is New Learning', 'Created By', 'Created At (MST)'
    ]
    ws.append(headers)
    
    # Data rows
    for project in queryset:
        serialized = ProjectSerializer(project).data
        row = [
            project.application_number,
            project.account_name,
            project.project_court,
            project.reviewed_by,
            project.project_status,
            project.stage,
            project.completed_date.strftime('%Y-%m-%d') if project.completed_date else '',
            serialized.get('start_time', ''),
            serialized.get('end_time', ''),
            float(project.total_time) if project.total_time else 0,
            project.partner_installer_account or '',
            project.third_party_salesforce or '',
            project.comments or '',
            project.content or '',
            'Yes' if project.is_new_learning else 'No',
            project.created_by.username,
            serialized.get('created_at', ''),
        ]
        ws.append(row)
    
    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=ie_logs_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    
    return response


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdmin])
def export_csv_view(request):
    """
    Export projects to CSV (admin only).
    Supports date range filtering and exports all columns in MST timezone.
    """
    # Get filter parameters
    start_date = request.data.get('start_date')
    end_date = request.data.get('end_date')
    project_ids = request.data.get('project_ids', [])
    
    # Build queryset
    queryset = Project.objects.filter(is_deleted=False).select_related('created_by')
    
    if start_date:
        queryset = queryset.filter(completed_date__gte=start_date)
    if end_date:
        queryset = queryset.filter(completed_date__lte=end_date)
    if project_ids:
        queryset = queryset.filter(id__in=project_ids)
    
    # Create CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename=ie_logs_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    
    writer = csv.writer(response)
    
    # Headers
    headers = [
        'Application #', 'Account Name', 'Project Court', 'Reviewed By',
        'Status', 'Stage', 'Completed Date', 'Start Time (MST)', 'End Time (MST)',
        'Total Time (hrs)', 'Partner Installer Account', 'Third Party Salesforce',
        'Comments', 'Content', 'Is New Learning', 'Created By', 'Created At (MST)'
    ]
    writer.writerow(headers)
    
    # Data rows
    for project in queryset:
        serialized = ProjectSerializer(project).data
        row = [
            project.application_number,
            project.account_name,
            project.project_court,
            project.reviewed_by,
            project.project_status,
            project.stage,
            project.completed_date.strftime('%Y-%m-%d') if project.completed_date else '',
            serialized.get('start_time', ''),
            serialized.get('end_time', ''),
            float(project.total_time) if project.total_time else 0,
            project.partner_installer_account or '',
            project.third_party_salesforce or '',
            project.comments or '',
            project.content or '',
            'Yes' if project.is_new_learning else 'No',
            project.created_by.username,
            serialized.get('created_at', ''),
        ]
        writer.writerow(row)
    
    return response
